from openpyxl import load_workbook, Workbook
from haversine import haversine
import heapq
import time


def excel_to_dict(excel_path, headers=[]):
    wb = load_workbook(excel_path)
    sheet = wb.worksheets[0]

    dict_list = []
    total = sheet.max_row
    for i, row in enumerate(sheet.rows):
        row_values = [c.value for idx, c in enumerate(row)]
        if i == 0:
            headers = row_values
        else:
            one_dict = dict(zip(headers, row_values))
            dict_list.append(one_dict)

        if (i + 1) % 100 == 0 or i + 1 == total:
            print('excel to dict: [{}/{}]'.format(i + 1, total))

    return dict_list


def find_shortest_straight(location_standard, locations_hospital, num_find=3):
    # 우선순위 (거리, index) 넣고 맨 위에 n개 빼서 리턴
    heap = []
    total = len(locations_hospital)
    for i, hospital in enumerate(locations_hospital):
        if not hospital['X']:
            if print_None:
                print('wrong!', hospital)
            continue
        distance = haversine((location_standard['Y'], location_standard['X']),
                             (hospital['Y'], hospital['X']))
        hospital['straight_distance'] = distance
        heapq.heappush(heap, (distance, i))

        # if (i + 1) % 100 == 0 or i + 1 == total:
        #     print('find shortest [{}/{}]'.format(i + 1, total))

    shortest = [heapq.heappop(heap)[1] for i in range(num_find)]
    return shortest


if __name__ == '__main__':
    start_time = time.time()

    # set hyper parameters
    path_to_offices = './resources/location_offices.xlsx'
    path_to_hospitals = './resources/hira_all_addr_w_coord.xlsx'

    all_hospitals = ['hira_subcategory_가정의학과.xlsx']
    dirname = './resources/'
    num_find = 3

    for hosname in all_hospitals:
        path_to_hospitals = dirname + hosname
        print(path_to_hospitals)

        out_name = 'shortest_' + path_to_hospitals.split('_')[-1]
        print(out_name)

        # load excel file and convert to list of dictionary
        offices = excel_to_dict(path_to_offices)
        hospitals = excel_to_dict(path_to_hospitals)

        # set workbook to save
        wb = Workbook()
        ws = wb.active

        len_offices = len(offices)
        for i, office in enumerate(offices):
            print_None = False
            if i == 1:
                print_None = True
            shortest = find_shortest_straight(office, hospitals, num_find)
            for num in range(num_find):
                office['OBJECTID'] = i + 1
                office['hname' + str(num + 1)] = hospitals[shortest[num]]['병원약국명']
                office['htype' + str(num + 1)] = hospitals[shortest[num]]['htype']
                office['Y_hosp' + str(num + 1)] = hospitals[shortest[num]]['Y']
                office['X_hosp' + str(num + 1)] = hospitals[shortest[num]]['X']
                office['dist_straight' + str(num + 1)] = hospitals[shortest[num]][
                    'straight_distance']

            if i == 0:
                ws.append(list(office.keys()))

            if (i + 1) % 20 == 0 or i + 1 == len_offices:
                print('find offices [{}/{}]'.format(i + 1, len_offices))
                wb.save(out_name)

            ws.append(list(office.values()))

            wb.save(out_name)
        print('saved to :', out_name)
        print('')
    elapsed_time = time.time() - start_time
    print('Took {:.03f} seconds'.format(elapsed_time))
