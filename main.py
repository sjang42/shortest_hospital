from openpyxl import load_workbook, Workbook
from haversine import haversine
import heapq
import time


def excel_to_dict(excel_path, headers=[]):
    wb = load_workbook(excel_path)
    sheet = wb.worksheets[0]

    dict_list = []
    total = sheet.max_row
    for i, row in enumerate(sheet.rows):
        row_values = [c.value for idx, c in enumerate(row)]
        if i == 0:
            headers = row_values
        else:
            one_dict = dict(zip(headers, row_values))
            dict_list.append(one_dict)

        if (i + 1) % 100 == 0 or i + 1 == total:
            print('excel to dict: [{}/{}]'.format(i + 1, total))

    return dict_list


def find_shortest_straight(location_standard, locations_hospital, num_find=3):
    # 우선순위 (거리, index) 넣고 맨 위에 n개 빼서 리턴
    heap = []
    total = len(locations_hospital)
    for i, hospital in enumerate(locations_hospital):
        distance = haversine((location_standard['Y'], location_standard['X']),
                             (hospital['Y'], hospital['X']))
        hospital['straight_distance'] = distance
        heapq.heappush(heap, (distance, i))

        # if (i + 1) % 100 == 0 or i + 1 == total:
        #     print('find shortest [{}/{}]'.format(i + 1, total))

    shortest = [heapq.heappop(heap)[1] for i in range(num_find)]
    return shortest

