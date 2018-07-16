from openpyxl import load_workbook


def excel_to_dict(excel_path, mute=False):
    """ Read an excel file and save as list - one dict per row
    :param excel_path: path to excel file
    :param mute: If it's true, dist func doesn't print logs
    :return: list of dict
    """
    wb = load_workbook(excel_path)
    sheet = wb.worksheets[0]

    dict_list = []
    total = sheet.max_row
    for i, row in enumerate(sheet.rows):
        row_values = [c.value for idx, c in enumerate(row)]
        if i == 0:
            headers = row_values
        else:
            one_dict = dict(zip(headers, row_values))
            dict_list.append(one_dict)

        if not mute and ((i + 1) % 500 == 0 or i + 1 == total):
            print('excel to dict: [{}/{}]'.format(i + 1, total))

    return dict_list


def is_excel(fname: str):
    """ Check if fname ends with excel extension
    :param fname: file name to check
    :return: Bool
    """

    ret = False
    if fname.endswith('.xlsx'):
        ret = True

    return ret


def tuple2str(tp):
    ret = ''

    for i, item in enumerate(tp):
        if i != 0:
            ret += ','
        ret += str(item)

    return ret


def list2str(iterable):
    ret = ''

    for i, item in enumerate(iterable):
        if i != 0:
            ret += ','
        ret += str(item)

    return ret


def float_trim(origin, limit=4):
    ret = ("{0:." + str(limit) + "f}").format(origin)

    return float(ret)
